import json
import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db import IntegrityError
from .models import Task, FocusSession, Profile, PlanItem

# --- STREAK HELPER ---
def update_user_streak(user):
    from django.utils import timezone
    import datetime
    profile, created = Profile.objects.get_or_create(user=user)
    
    sessions = FocusSession.objects.filter(user=user).order_by('-completed_at')
    if not sessions.exists():
        profile.streak = 0
        profile.save(update_fields=['streak'])
        return 0
        
    local_dates = set()
    for s in sessions:
        local_date = timezone.localtime(s.completed_at).date()
        local_dates.add(local_date)
        
    sorted_dates = sorted(list(local_dates), reverse=True)
    today = timezone.localtime(timezone.now()).date()
    yesterday = today - datetime.timedelta(days=1)
    
    if today not in sorted_dates and yesterday not in sorted_dates:
        profile.streak = 0
        profile.save(update_fields=['streak'])
        return 0
        
    streak = 0
    current_expected = sorted_dates[0]
    
    for date in sorted_dates:
        if date == current_expected:
            streak += 1
            current_expected -= datetime.timedelta(days=1)
        elif date < current_expected:
            break
            
    profile.streak = streak
    if sorted_dates:
        profile.last_completion_date = sorted_dates[0]
    profile.save(update_fields=['streak', 'last_completion_date'])
    return streak

# --- DASHBOARD HOME ---
@login_required
def home(request):
    # Recalculate streak dynamically
    update_user_streak(request.user)
    # Retrieve user-specific profile
    profile, created = Profile.objects.get_or_create(user=request.user)
    
    # Retrieve user-specific tasks, sessions, and plan checklist items
    tasks = Task.objects.filter(user=request.user).order_by('-created_at')
    sessions = FocusSession.objects.filter(user=request.user).order_by('-completed_at')
    plan_items = PlanItem.objects.filter(user=request.user).order_by('order')

    if request.method == "POST":
        title = request.POST.get("title")
        duration = request.POST.get("duration")
        rules = request.POST.get("rules")
        assigned_date = request.POST.get("assigned_date")

        if not assigned_date:
            assigned_date = timezone.now().date()

        Task.objects.create(
            user=request.user,
            title=title,
            duration=duration,
            rules=rules,
            assigned_date=assigned_date
        )
        return redirect('home')

    context = {
        'tasks': tasks,
        'sessions': sessions,
        'plan_items': plan_items,
        'profile': profile,
    }
    return render(request, 'index.html', context)


# --- TASK OPERATIONS & GAMIFICATION SYSTEM ---
@login_required
def complete_task(request, id):
    task = get_object_or_404(Task, id=id, user=request.user)
    task.completed = True
    task.completed_at = timezone.now()
    task.save()

    # Automatically log a focus session in the database
    FocusSession.objects.create(
        user=request.user,
        task_title=task.title,
        duration=task.duration,
        completed_at=timezone.now()
    )

    # Award XP and update Streaks/Badges
    profile, created = Profile.objects.get_or_create(user=request.user)
    
    # 100 base + duration * 5 XP
    earned_xp = 100 + (task.duration * 5)
    profile.xp += earned_xp
    profile.save(update_fields=['xp'])

    # Recalculate streak dynamically based on the updated focus sessions
    update_user_streak(request.user)
    
    # Reload profile to get the updated streak and last completion date
    profile.refresh_from_db()

    # Check and Unlock Badges
    try:
        badges = json.loads(profile.badges)
    except Exception:
        badges = []

    completed_count = Task.objects.filter(user=request.user, completed=True).count()
    
    if completed_count >= 1 and 'badge-initiate' not in badges:
        badges.append('badge-initiate')
    if profile.xp >= 500 and 'badge-runner' not in badges:
        badges.append('badge-runner')
    if task.duration >= 60 and 'badge-master' not in badges:
        badges.append('badge-master')
    if profile.streak >= 3 and 'badge-streak' not in badges:
        badges.append('badge-streak')
    if profile.xp >= 2000 and 'badge-legend' not in badges:
        badges.append('badge-legend')
    if completed_count >= 5 and 'badge-grind' not in badges:
        badges.append('badge-grind')

    profile.badges = json.dumps(badges)
    profile.save(update_fields=['badges'])

    messages.success(request, f"Task complete! Earned +{earned_xp} XP!")
    return redirect('home')


@login_required
def edit_task(request, id):
    task = get_object_or_404(Task, id=id, user=request.user)
    if request.method == "POST":
        task.title = request.POST.get("title")
        task.duration = request.POST.get("duration")
        task.rules = request.POST.get("rules")
        assigned_date = request.POST.get("assigned_date")
        if assigned_date:
            task.assigned_date = assigned_date
        task.save()
        messages.success(request, "Task updated successfully!")
    return redirect('home')


@login_required
def delete_task(request, id):
    task = get_object_or_404(Task, id=id, user=request.user)
    task.delete()
    messages.success(request, "Task deleted.")
    return redirect('home')


# --- AUTHENTICATION ---
def signup_view(request):
    if request.user.is_authenticated:
        return redirect('home')
        
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')

        if not (username and email and password and confirm_password):
            messages.error(request, "All fields are required.")
            return render(request, 'signup.html')

        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return render(request, 'signup.html')

        if len(password) < 6:
            messages.error(request, "Password must be at least 6 characters.")
            return render(request, 'signup.html')

        try:
            user = User.objects.create_user(username=username, email=email, password=password)
            # Log the user in directly
            login(request, user)
            messages.success(request, f"Welcome to Velora, {username}!")
            return redirect('home')
        except IntegrityError:
            messages.error(request, "Username is already taken.")
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")

    return render(request, 'signup.html')


def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        username_or_email = request.POST.get('username_or_email', '').strip()
        password = request.POST.get('password')
        remember_me = request.POST.get('remember_me')

        if not (username_or_email and password):
            messages.error(request, "Please fill in all fields.")
            return render(request, 'login.html')

        # Find user by username or email
        username = username_or_email
        if '@' in username_or_email:
            try:
                user_obj = User.objects.get(email=username_or_email)
                username = user_obj.username
            except User.DoesNotExist:
                pass

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            if remember_me:
                # 2 weeks
                request.session.set_expiry(1209600)
            else:
                # browser close
                request.session.set_expiry(0)
            messages.success(request, f"Welcome back, {user.username}!")
            return redirect('home')
        else:
            messages.error(request, "Invalid username/email or password.")

    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    messages.success(request, "You have been logged out.")
    return redirect('login')


# --- PROFILE AND SETTINGS ---
@login_required
def profile_view(request):
    # Recalculate streak dynamically
    update_user_streak(request.user)

    profile, created = Profile.objects.get_or_create(user=request.user)
    
    # Calculate statistics
    completed_tasks_count = Task.objects.filter(user=request.user, completed=True).count()
    sessions = FocusSession.objects.filter(user=request.user)
    total_duration_mins = sum(s.duration for s in sessions)
    total_hours = round(total_duration_mins / 60, 1)

    try:
        badges = json.loads(profile.badges)
    except Exception:
        badges = []

    badge_meta = {
        'badge-initiate': {'name': 'Initiate', 'icon': 'award', 'desc': 'Complete your 1st task'},
        'badge-runner': {'name': 'Runner', 'icon': 'rocket', 'desc': 'Earn 500 XP'},
        'badge-master': {'name': 'Master', 'icon': 'shield', 'desc': 'Complete a 60m+ task'},
        'badge-streak': {'name': 'Aligned', 'icon': 'zap', 'desc': 'Reach a 3-day streak'},
        'badge-legend': {'name': 'Legend', 'icon': 'crown', 'desc': 'Earn 2,000 XP'},
        'badge-grind': {'name': 'Grind', 'icon': 'hourglass', 'desc': 'Complete 5 tasks'},
    }

    unlocked_badges_list = []
    for b_id in badges:
        if b_id in badge_meta:
            unlocked_badges_list.append(badge_meta[b_id])

    context = {
        'profile': profile,
        'completed_tasks_count': completed_tasks_count,
        'total_hours': total_hours,
        'badges_unlocked': unlocked_badges_list,
    }
    return render(request, 'profile.html', context)


@login_required
def settings_view(request):
    profile, created = Profile.objects.get_or_create(user=request.user)
    user = request.user

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'personal':
            username = request.POST.get('username', '').strip()
            email = request.POST.get('email', '').strip()
            
            if username and username != user.username:
                if User.objects.filter(username=username).exists():
                    messages.error(request, "Username is already taken.")
                else:
                    user.username = username
                    user.save()
            
            if email and email != user.email:
                user.email = email
                user.save()
            
            messages.success(request, "Personal settings updated.")
            return redirect('settings')
            
        elif action == 'security':
            password = request.POST.get('password')
            confirm_password = request.POST.get('confirm_password')
            
            if password and password == confirm_password:
                if len(password) < 6:
                    messages.error(request, "Password must be at least 6 characters.")
                else:
                    user.set_password(password)
                    user.save()
                    login(request, user)  # Re-authenticate to prevent logout
                    messages.success(request, "Password updated successfully.")
                    return redirect('settings')
            else:
                messages.error(request, "Passwords do not match.")
                
        elif action == 'avatar':
            avatar_file = request.FILES.get('avatar')
            if avatar_file:
                profile.avatar = avatar_file
                profile.save()
                messages.success(request, "Avatar updated successfully.")
                return redirect('settings')
            else:
                messages.error(request, "No file uploaded.")

    return render(request, 'settings.html', {'profile': profile})


@login_required
def settings_theme_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            theme = data.get('theme')
            if theme in ['violet', 'dark', 'light']:
                profile, created = Profile.objects.get_or_create(user=request.user)
                profile.theme_preference = theme
                profile.save()
                return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)


# --- TODAY'S PLAN CHECKLIST AJAX APIs ---
@login_required
def plan_add(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            text = data.get('text', '').strip()
            if not text:
                return JsonResponse({'status': 'error', 'message': 'Text is empty'}, status=400)

            # Determine order (max order + 1)
            max_order_item = PlanItem.objects.filter(user=request.user).order_by('-order').first()
            new_order = (max_order_item.order + 1) if max_order_item else 0

            item = PlanItem.objects.create(
                user=request.user,
                text=text,
                completed=False,
                order=new_order
            )
            return JsonResponse({
                'status': 'success',
                'id': item.id,
                'text': item.text,
                'completed': item.completed
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)


@login_required
def plan_toggle(request, id):
    if request.method == 'POST':
        item = get_object_or_404(PlanItem, id=id, user=request.user)
        item.completed = not item.completed
        item.save()
        return JsonResponse({
            'status': 'success',
            'id': item.id,
            'completed': item.completed
        })
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)


@login_required
def plan_delete(request, id):
    if request.method == 'POST':
        item = get_object_or_404(PlanItem, id=id, user=request.user)
        item.delete()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)


@login_required
def plan_reorder(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            item_ids = data.get('item_ids', [])
            
            for idx, item_id in enumerate(item_ids):
                PlanItem.objects.filter(id=item_id, user=request.user).update(order=idx)
                
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)


# --- EXPORT REPORTING UTILITIES ---
@login_required
def export_tasks_pdf(request):
    # Generates a premium PDF report of user's task configurations and schedule
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Velora_Tasks_{timezone.now().strftime("%Y%m%d")}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    story = []

    # Palette variables (Match Violet Theme styling)
    primary_color = colors.HexColor('#7c3aed') # Violet
    dark_bg = colors.HexColor('#0c0418')
    text_color = colors.HexColor('#1f2937')
    light_purple = colors.HexColor('#f5f3ff')

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=primary_color,
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#6b7280'),
        spaceAfter=30
    )
    th_style = ParagraphStyle(
        'TableHeader',
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.white
    )
    tb_style = ParagraphStyle(
        'TableBody',
        fontName='Helvetica',
        fontSize=9,
        textColor=text_color
    )

    # Header
    story.append(Paragraph("Velora Scheduler Report", title_style))
    story.append(Paragraph(f"Generated for: {request.user.username}  |  Date: {timezone.now().strftime('%B %d, %Y')}", subtitle_style))
    story.append(Spacer(1, 10))

    # Fetch tasks
    tasks = Task.objects.filter(user=request.user).order_by('-assigned_date', '-created_at')
    
    # Table Header
    data = [
        [
            Paragraph("Task Description", th_style),
            Paragraph("Assigned Date", th_style),
            Paragraph("Mins", th_style),
            Paragraph("Status", th_style),
            Paragraph("Rules/Guidelines", th_style)
        ]
    ]

    # Table Data
    for t in tasks:
        status = "Completed" if t.completed else "Pending"
        data.append([
            Paragraph(t.title, tb_style),
            Paragraph(t.assigned_date.strftime("%Y-%m-%d"), tb_style),
            Paragraph(str(t.duration), tb_style),
            Paragraph(status, ParagraphStyle('status', parent=tb_style, textColor=colors.HexColor('#16a34a') if t.completed else primary_color)),
            Paragraph(t.rules or "-", tb_style)
        ])

    table = Table(data, colWidths=[130, 80, 45, 65, 210])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), primary_color),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_purple]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e5e7eb')),
        ('TOPPADDING', (0,1), (-1,-1), 6),
        ('BOTTOMPADDING', (0,1), (-1,-1), 6),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))

    story.append(table)
    doc.build(story)
    return response


@login_required
def export_history_pdf(request):
    # Generates a premium PDF report of focus session history
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Velora_FocusHistory_{timezone.now().strftime("%Y%m%d")}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    story = []

    primary_color = colors.HexColor('#7c3aed')
    light_purple = colors.HexColor('#f5f3ff')
    text_color = colors.HexColor('#1f2937')

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=primary_color,
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#6b7280'),
        spaceAfter=30
    )
    th_style = ParagraphStyle(
        'TableHeader',
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.white
    )
    tb_style = ParagraphStyle(
        'TableBody',
        fontName='Helvetica',
        fontSize=9,
        textColor=text_color
    )

    story.append(Paragraph("Velora Focus Session Logs", title_style))
    story.append(Paragraph(f"Generated for: {request.user.username}  |  Date: {timezone.now().strftime('%B %d, %Y')}", subtitle_style))
    story.append(Spacer(1, 10))

    sessions = FocusSession.objects.filter(user=request.user).order_by('-completed_at')

    # Calculate Totals
    total_mins = sum(s.duration for s in sessions)
    total_hrs = round(total_mins / 60, 1)

    stat_summary_style = ParagraphStyle(
        'SummaryStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=11,
        textColor=text_color,
        spaceAfter=20
    )
    story.append(Paragraph(f"Summary Metrics: Total Logged Sessions: {len(sessions)}  |  Accumulated Hours: {total_hrs} Hours", stat_summary_style))

    data = [
        [
            Paragraph("Task Description / Title", th_style),
            Paragraph("Focused Duration (minutes)", th_style),
            Paragraph("Completed Timestamp (UTC)", th_style)
        ]
    ]

    for s in sessions:
        data.append([
            Paragraph(s.task_title, tb_style),
            Paragraph(f"{s.duration} mins", tb_style),
            Paragraph(s.completed_at.strftime("%Y-%m-%d %H:%M:%S"), tb_style)
        ])

    table = Table(data, colWidths=[240, 130, 160])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), primary_color),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_purple]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e5e7eb')),
        ('TOPPADDING', (0,1), (-1,-1), 6),
        ('BOTTOMPADDING', (0,1), (-1,-1), 6),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))

    story.append(table)
    doc.build(story)
    return response


@login_required
def export_stats_excel(request):
    # Generates a premium multi-tab Excel spreadsheet containing overview, tasks list, and focus session list
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Velora_Analytics_{timezone.now().strftime("%Y%m%d")}.xlsx"'

    wb = Workbook()

    # Sheet 1: Dashboard Overview & Statistics
    ws1 = wb.active
    ws1.title = "Overview Metrics"
    ws1.views.sheetView[0].showGridLines = True

    # Styling colors
    primary_color = "7C3AED"  # Purple
    light_purple_fill = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")
    header_fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color=primary_color)
    meta_font = Font(name="Calibri", size=10, italic=True, color="6B7280")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)

    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    # Title & Metadata
    ws1['A1'] = "Velora Productivity Overview"
    ws1['A1'].font = title_font
    ws1['A2'] = f"Generated for user: {request.user.username}  |  Timestamp: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws1['A2'].font = meta_font

    # Fetch raw data
    update_user_streak(request.user)
    tasks = Task.objects.filter(user=request.user)
    sessions = FocusSession.objects.filter(user=request.user)
    profile, created = Profile.objects.get_or_create(user=request.user)

    total_tasks = tasks.count()
    completed_tasks = tasks.filter(completed=True).count()
    pending_tasks = total_tasks - completed_tasks
    completion_rate = round((completed_tasks / total_tasks * 100), 1) if total_tasks > 0 else 0.0

    total_focus_mins = sum(s.duration for s in sessions)
    total_focus_hours = round(total_focus_mins / 60, 2)
    longest_session = max((s.duration for s in sessions), default=0)

    # Overview Table
    ws1['A4'] = "Metric Item"
    ws1['B4'] = "Calculated Value"
    ws1['A4'].fill = header_fill
    ws1['A4'].font = header_font
    ws1['B4'].fill = header_fill
    ws1['B4'].font = header_font

    metrics = [
        ("Current User XP Level", f"Level {int(profile.xp / 1000) + 1}"),
        ("Total Acquired XP", f"{profile.xp} XP"),
        ("Current Day Streak", f"{profile.streak} Days"),
        ("Total Scheduled Tasks", total_tasks),
        ("Completed Tasks", completed_tasks),
        ("Pending Tasks", pending_tasks),
        ("Task Completion Rate", f"{completion_rate}%"),
        ("Total Focus Hours", f"{total_focus_hours} Hours"),
        ("Longest Focus Session", f"{longest_session} Minutes"),
    ]

    for idx, (m_name, m_val) in enumerate(metrics, start=5):
        ws1[f'A{idx}'] = m_name
        ws1[f'B{idx}'] = m_val
        ws1[f'A{idx}'].font = normal_font
        ws1[f'B{idx}'].font = bold_font
        ws1[f'A{idx}'].border = thin_border
        ws1[f'B{idx}'].border = thin_border
        if idx % 2 == 0:
            ws1[f'A{idx}'].fill = light_purple_fill
            ws1[f'B{idx}'].fill = light_purple_fill

    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 20

    # Sheet 2: Completed Focus Sessions
    ws2 = wb.create_sheet(title="Focus Sessions Log")
    ws2.views.sheetView[0].showGridLines = True
    ws2['A1'] = "Logged Focus Sessions"
    ws2['A1'].font = title_font
    
    headers_s = ["Focus Session ID", "Task Title Description", "Session Duration (mins)", "Completed Timestamp (UTC)"]
    for col_idx, h in enumerate(headers_s, start=1):
        cell = ws2.cell(row=3, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill

    for r_idx, s in enumerate(sessions, start=4):
        ws2.cell(row=r_idx, column=1, value=s.id).font = normal_font
        ws2.cell(row=r_idx, column=2, value=s.task_title).font = normal_font
        ws2.cell(row=r_idx, column=3, value=s.duration).font = normal_font
        ws2.cell(row=r_idx, column=4, value=s.completed_at.strftime("%Y-%m-%d %H:%M:%S")).font = normal_font
        for col_idx in range(1, 5):
            ws2.cell(row=r_idx, column=col_idx).border = thin_border
            if r_idx % 2 == 0:
                ws2.cell(row=r_idx, column=col_idx).fill = light_purple_fill

    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 22
    ws2.column_dimensions['D'].width = 25

    # Sheet 3: Tasks List Details
    ws3 = wb.create_sheet(title="Tasks Records")
    ws3.views.sheetView[0].showGridLines = True
    ws3['A1'] = "Scheduled Tasks Directory"
    ws3['A1'].font = title_font

    headers_t = ["Task ID", "Task Title", "Session Mins", "Status", "Assigned Date", "Created At", "Completed At"]
    for col_idx, h in enumerate(headers_t, start=1):
        cell = ws3.cell(row=3, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill

    for r_idx, t in enumerate(tasks, start=4):
        status = "Completed" if t.completed else "Pending"
        c_at = t.completed_at.strftime("%Y-%m-%d %H:%M:%S") if t.completed_at else "-"
        
        ws3.cell(row=r_idx, column=1, value=t.id).font = normal_font
        ws3.cell(row=r_idx, column=2, value=t.title).font = normal_font
        ws3.cell(row=r_idx, column=3, value=t.duration).font = normal_font
        ws3.cell(row=r_idx, column=4, value=status).font = bold_font if t.completed else normal_font
        ws3.cell(row=r_idx, column=5, value=t.assigned_date.strftime("%Y-%m-%d")).font = normal_font
        ws3.cell(row=r_idx, column=6, value=t.created_at.strftime("%Y-%m-%d %H:%M:%S")).font = normal_font
        ws3.cell(row=r_idx, column=7, value=c_at).font = normal_font
        
        for col_idx in range(1, 8):
            ws3.cell(row=r_idx, column=col_idx).border = thin_border
            if r_idx % 2 == 0:
                ws3.cell(row=r_idx, column=col_idx).fill = light_purple_fill

    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 28
    ws3.column_dimensions['C'].width = 14
    ws3.column_dimensions['D'].width = 12
    ws3.column_dimensions['E'].width = 15
    ws3.column_dimensions['F'].width = 22
    ws3.column_dimensions['G'].width = 22

    wb.save(response)
    return response